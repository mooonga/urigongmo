from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from datetime import datetime
from poster.models import Poster

class Command(BaseCommand):
    help = "Excel에서 포스터 데이터를 불러와 DB에 저장합니다."

    def handle(self, *args, **kwargs):
        wb = load_workbook('data/poster.xlsx')
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            title = row[0]
            d_day = row[1]
            image_filename = row[2]  # 예: poster_2.png
            organization = row[3]
            company_type = row[4]
            target = row[5]
            prize = row[6]
            start_date = row[7]
            end_date = row[8]
            website = row[9]
            benefits = row[10]
            category = row[11]
            category_id = row[12]
            extra = row[13]
            description = row[14]

            # 날짜 문자열이면 변환
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y.%m.%d').date()
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y.%m.%d').date()

            Poster.objects.create(
                title=title,
                d_day=d_day,
                image=f'poster_images/{image_filename}',
                organization=organization,
                company_type=company_type,
                target=target,
                prize=prize,
                start_date=start_date,
                end_date=end_date,
                website=website,
                benefits=benefits,
                category=category,
                category_id=int(category_id) if category_id else None,
                extra=extra,
                description=description,
            )

        self.stdout.write(self.style.SUCCESS("✅ 이미지 경로 포함 포스터 DB 등록 완료!"))
