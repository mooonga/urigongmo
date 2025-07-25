from openpyxl import load_workbook
from PIL import Image
import io
import os
import sys
import django

# Django 환경 설정
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + '/../')  # config 폴더 상위까지 경로 지정
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from poster.models import Poster  # 모델 import

# 엑셀 파일 열기
wb = load_workbook('data/poster.xlsx')
ws = wb.active

# 이미지 저장 경로
save_dir = 'media/poster_images/'
os.makedirs(save_dir, exist_ok=True)

# 이미지 객체들을 반복
for idx, image in enumerate(ws._images):
    # 이미지가 속한 셀 (예: 'F2')
    cell_ref = image.anchor._from.row + 1  # 0부터 시작이라 +1
    image_name = f"poster_{cell_ref}.png"
    image_path = os.path.join(save_dir, image_name)

    # 이미지 저장
    img = Image.open(io.BytesIO(image._data()))
    img.save(image_path)

    # 엑셀 데이터 읽기
    title = ws[f'B{cell_ref}'].value
    year = str(ws[f'C{cell_ref}'].value)
    description = ws[f'E{cell_ref}'].value  # ← 설명은 E열에서 추출
    print(f"[{cell_ref}] {title} / 설명: {description}")  # 확인용 출력
    
    # DB에 저장
    Poster.objects.create(
        title=title,
        year=year,
        image=f'poster_images/{image_name}',
        description=description
    )

print("✅ 이미지 저장 및 DB 등록 완료!")
