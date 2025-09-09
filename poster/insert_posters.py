from openpyxl import load_workbook
from PIL import Image
from datetime import datetime
import io
import os
import django
import sys

# Django 환경 설정
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + '/../')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from poster.models import Poster

# 엑셀 파일 열기
wb = load_workbook('data/poster.xlsx')
ws = wb.active

# 이미지 저장 경로
save_dir = 'media/poster_images/'
os.makedirs(save_dir, exist_ok=True)

# 이미지 처리 + DB 저장
for idx, image in enumerate(ws._images):
    cell_row = image.anchor._from.row + 1  # 이미지가 있는 행
    image_name = f"poster_{cell_row}.png"
    image_path = os.path.join(save_dir, image_name)

    # 이미지 저장
    img = Image.open(io.BytesIO(image._data()))
    img.save(image_path)

    # 엑셀에서 데이터 읽기
    title = ws[f'B{cell_row}'].value
    year = str(ws[f'C{cell_row}'].value)
    description = ws[f'E{cell_row}'].value  # 범위와 내용
    start_date = datetime.strptime(ws[f'H{cell_row}'].value, "%Y.%m.%d").date()
    end_date = datetime.strptime(ws[f'I{cell_row}'].value, "%Y.%m.%d").date()

    # 모델에 저장
    Poster.objects.create(
        title=title,
        image=f'poster_images/{image_name}',
        start_date=start_date,
        end_date=end_date,
        description=description
    )

print("포스터 등록 완료!")
